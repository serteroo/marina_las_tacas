import random, datetime
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.core.mail import send_mail
from django.contrib import messages
from .models import MFAChallenge, UserProfile, Organization, ContratoExterno
from .forms import ContratoExternoForm
from datetime import timedelta 
from operaciones.models import Embarcacion, Movimiento, BloqueoClima
from django.conf import settings
from django.views.decorators.cache import never_cache
from django.utils.decorators import method_decorator
from django.contrib.auth.hashers import make_password
from .forms import PublicRegisterForm
from .models import Applicant, Profile
import secrets
from django.contrib.auth.views import PasswordChangeView, PasswordChangeDoneView
from django.urls import reverse_lazy, reverse
from django.apps import apps
from django.utils.crypto import get_random_string
from django.db import transaction, models
from django.db.models import Q
from django.contrib.messages import get_messages
from django import forms
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl
from django.core.paginator import Paginator
from django.contrib.auth.models import Group
from .roles import is_supervisor, is_supervisor_strict, is_secretaria, is_socio, menu_kind_for
from .security import get_client_ip, is_login_blocked, register_login_attempt




def _send_mfa_email(user, code):
    send_mail(
        subject="Tu código de verificación",
        message=f"Tu código es: {code}",
        from_email=None, recipient_list=[user.email], fail_silently=False
    )

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")

        # NUEVO: IP del cliente
        ip = get_client_ip(request)

        # NUEVO: bloqueo por demasiados intentos fallidos
        if is_login_blocked(username, ip):
            messages.error(
                request,
                "Demasiados intentos fallidos de inicio de sesión. Intenta más tarde."
            )
            return render(request, "registration/login.html", {"blocked": True})

        # Paso 1: autenticar usuario (solo password, como ya hacías)
        user = authenticate(request, username=username, password=password)

        if user:
            # ✅ Login de password correcto → registrar éxito y seguir con tu MFA

            register_login_attempt(username, ip, success=True)

            # perfil obligatorio
            try:
                profile = UserProfile.objects.get(user=user)
            except UserProfile.DoesNotExist:
                messages.warning(request, "No existe perfil de usuario. Contacta al administrador.")
                return render(request, "registration/login.html")

            # backend para luego loguear tras MFA
            backend = getattr(user, "backend", None) or settings.AUTHENTICATION_BACKENDS[0]

            # MFA OBLIGATORIO
            if not user.email:
                messages.error(request, "Tu usuario no tiene email configurado. Contacta al administrador.")
                return render(request, "registration/login.html")

            # opcional: limpia desafíos previos aún vigentes
            MFAChallenge.objects.filter(
                user=user,
                used=False,
                expires_at__gt=timezone.now()
            ).delete()

            code = f"{random.randint(0, 999999):06d}"
            exp = timezone.now() + datetime.timedelta(minutes=5)
            MFAChallenge.objects.create(
                user=user,
                code=code,
                channel="email",
                expires_at=exp
            )
            _send_mfa_email(user, code)

            if settings.DEBUG:
                request.session["dev_mfa_code"] = code  # solo para DEBUG

            request.session["pending_user_id"] = user.id
            request.session["pending_auth_backend"] = backend

            # guarda next para usarlo tras completar 2FA
            next_url = request.GET.get("next") or request.POST.get("next")
            if next_url:
                request.session["post_login_redirect"] = next_url

            messages.info(request, "Te enviamos un código de verificación al correo.")
            return redirect("accounts:two_factor")

        # ❌ Credenciales inválidas → registrar intento fallido y responder
        register_login_attempt(username, ip, success=False)
        messages.error(request, "Credenciales inválidas.")
        return render(request, "registration/login.html")

    # GET: limpia mensajes y SIEMPRE devuelve el template
    list(get_messages(request))
    return render(request, "registration/login.html")


RESEND_COOLDOWN_SECONDS = 30

def _seconds_to_resend(user_id: int) -> int:
    """Devuelve cuantos segundos faltan para poder reenviar (0 si ya se puede)."""
    last = (MFAChallenge.objects
            .filter(user_id=user_id, used=False)
            .order_by("-created_at")
            .first())
    if not last:
        return 0
    elapsed = (timezone.now() - last.created_at).total_seconds()
    remaining = RESEND_COOLDOWN_SECONDS - int(elapsed)
    return remaining if remaining > 0 else 0


def two_factor_view(request):
    pending_id = request.session.get("pending_user_id")
    if not pending_id:
        messages.info(request, "Sesión 2FA no encontrada. Inicia sesión nuevamente.")
        return redirect("accounts:login")

    if request.method == "POST":
        code = request.POST.get("code", "").strip()

        try:
            ch = (MFAChallenge.objects
                  .filter(user_id=pending_id, used=False)
                  .latest("created_at"))
        except MFAChallenge.DoesNotExist:
            messages.error(request, "No hay desafío activo. Inicia sesión nuevamente.")
            return redirect("accounts:login")

        if ch.expires_at < timezone.now():
            messages.error(request, "Código expirado. Inicia sesión otra vez.")
            return redirect("accounts:login")

        if ch.code != code:
            messages.error(request, "Código incorrecto.")
            # volvemos a pintar, pero enviando los segundos restantes de cooldown
            return render(request, "registration/two_factor.html", {
                "resend_seconds": _seconds_to_resend(pending_id),
            })

        ch.used = True
        ch.save(update_fields=["used"])

        UserModel = get_user_model()
        user = UserModel.objects.get(id=pending_id)
        backend = request.session.pop("pending_auth_backend", None) or settings.AUTHENTICATION_BACKENDS[0]
        login(request, user, backend=backend)
        request.session.pop("pending_user_id", None)

        if hasattr(user, "profile") and getattr(user.profile, "must_change_password", False):
            return redirect("accounts:password_change")

        return redirect("accounts:dashboard")

    # GET
    return render(request, "registration/two_factor.html", {
        "resend_seconds": _seconds_to_resend(pending_id),
    })


def two_factor_resend(request):
    """Reenvía el código si pasó el cooldown."""
    pending_id = request.session.get("pending_user_id")
    if not pending_id:
        messages.info(request, "Sesión 2FA no encontrada. Inicia sesión nuevamente.")
        return redirect("accounts:login")

    remaining = _seconds_to_resend(pending_id)
    if remaining > 0:
        messages.info(request, f"Podrás reenviar en {remaining}s.")
        return redirect("accounts:two_factor")

    # Generar y enviar nuevo código
    UserModel = get_user_model()
    user = UserModel.objects.get(id=pending_id)

    code = f"{random.randint(0, 999999):06d}"
    exp = timezone.now() + timedelta(minutes=5)
    MFAChallenge.objects.create(user=user, code=code, channel="email", expires_at=exp)
    _send_mfa_email(user, code)

    messages.success(request, "Te enviamos un nuevo código a tu correo.")
    return redirect("accounts:two_factor")

@login_required
def dashboard(request):
    if is_supervisor(request.user):
        return redirect("accounts:dashboard_supervisor")
    return redirect("accounts:dashboard_socio")


# --- crear contrato: toma organización del perfil y fuerza licencia_validada=False ---
@login_required
@user_passes_test(is_supervisor)
def contrato_externo_new(request):
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    org = prof.organization

    if request.method == "POST":
        form = ContratoExternoForm(request.POST, request.FILES)
        if form.is_valid():
            fv = form.cleaned_data.get("licencia_vencimiento")
            hoy = timezone.localdate()
            if not fv or fv <= hoy:
                form.add_error("licencia_vencimiento", "Licencia vencida o no informada.")
                return render(request, "accounts/contrato_externo_form.html", {"form": form})

            c = form.save(commit=False)
            c.organization = org
            c.licencia_validada = True
            c.estado = "APROBADO"     # <- aprobado directo
            c.save()
            messages.success(request, "Contrato creado y aprobado.")
            return redirect("accounts:dashboard_supervisor")
    else:
        form = ContratoExternoForm()
    return render(request, "accounts/contrato_externo_form.html", {"form": form})


# --- detalle ---
@login_required
@user_passes_test(is_supervisor)
def contrato_externo_detail(request, pk):
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    contrato = get_object_or_404(ContratoExterno, pk=pk, organization=prof.organization)
    return render(request, "accounts/contrato_externo_detail.html", {"contrato": contrato})


# --- helper para saber si la licencia está vencida ---
def _licencia_vencida(fecha):
    if not fecha:
        return True
    return fecha < timezone.localdate()


# --- aprobar: valida licencia (número, validada y no vencida) ---
@login_required
@user_passes_test(is_supervisor)
def contrato_externo_aprobar(request, pk):
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    contrato = get_object_or_404(ContratoExterno, pk=pk, organization=prof.organization)
    errores = []

    if not contrato.licencia_numero:
        errores.append("El solicitante no registró número de licencia.")
    if not contrato.licencia_validada:
        errores.append("Debes marcar la licencia como validada antes de aprobar.")
    if _licencia_vencida(contrato.licencia_vencimiento):
        errores.append("La licencia está vencida o sin fecha de vencimiento.")

    if errores:
        for e in errores:
            messages.error(request, e)
        return redirect("accounts:contrato_externo_detail", pk=pk)

    contrato.estado = "APROBADO"
    contrato.supervisor = request.user
    contrato.save(update_fields=["estado", "supervisor", "actualizado_en"])
    messages.success(request, "Contrato aprobado.")
    return redirect("accounts:contrato_externo_detail", pk=pk)


# --- rechazar ---
@login_required
@user_passes_test(is_supervisor)
def contrato_externo_rechazar(request, pk):
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    contrato = get_object_or_404(ContratoExterno, pk=pk, organization=prof.organization)
    contrato.estado = "RECHAZADO"
    contrato.supervisor = request.user
    contrato.save(update_fields=["estado", "supervisor", "actualizado_en"])
    messages.warning(request, "Contrato rechazado.")
    return redirect("accounts:contrato_externo_detail", pk=pk)

@login_required
@user_passes_test(is_supervisor)
def dashboard_supervisor(request):
    # Organización del usuario
    request.session["active_dashboard"] = "supervisor"
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    org = prof.organization

    # --------- Métricas de movimientos (lo que ya tenías) ---------
    ctx = {
        "total_solicitados":  Movimiento.objects.filter(organization=org, estado="SOLICITADO").count(),
        "total_aprobados":    Movimiento.objects.filter(organization=org, estado="APROBADO").count(),
        "en_salida":          Movimiento.objects.filter(organization=org, estado="EN_SALIDA").count(),
        "pendientes_arribo":  Movimiento.objects.filter(organization=org, estado__in=["EN_SALIDA","EN_ARRIBO"]).count(),
        "ultimos":            Movimiento.objects.filter(organization=org)
                                .select_related("embarcacion","socio__user").order_by("-id")[:10],
        "bloqueo":            BloqueoClima.objects.filter(organization=org).order_by("-created_at","-id").first(),
        "pendientes":         Movimiento.objects.filter(organization=org, estado="SOLICITADO")
                                .select_related("embarcacion","socio__user").order_by("id"),
    }

    # Postulantes (lo que ya tenías)
    Applicant = apps.get_model('accounts', 'Applicant')
    pend_aplicants_qs = Applicant.objects.filter(estado='pending')
    ctx["postulantes_pendientes_count"] = pend_aplicants_qs.count()

    # --------- Contratos Externos (flujo supervisor crea=aprobado) ---------
    hoy = timezone.localdate()
    cx_base = ContratoExterno.objects.filter(organization=org)

    cx_registrados_hoy = cx_base.filter(
        estado="APROBADO", actualizado_en__date=hoy
    ).count()

    cx_ultimos = cx_base.order_by("-creado_en")[:5]   # <= tope duro a 5

    ctx.update({
        "cx_registrados_hoy": cx_registrados_hoy,
        "cx_ultimos": cx_ultimos,
        "menu_kind": menu_kind_for(request.user),
        "active_nav": "dashboard",
    })
    return render(request, "dashboard/supervisor.html", ctx)

@never_cache
@login_required
@user_passes_test(is_socio)
def dashboard_socio(request):
    request.session["active_dashboard"] = "socio"
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    org = prof.organization  # <<< clave

    # 2) embarcaciones del socio (clave)
    embarcaciones = (
        Embarcacion.objects
        .filter(organization=org, propietario=prof)  # <<< filtro por org
        .select_related("amarra", "organization", "propietario")
        .order_by("matricula")
    )

    # 3) últimos movimientos del socio
    ultimos = (
        Movimiento.objects
        .filter(organization=org, socio=prof)        # <<< filtro por org
        .select_related("embarcacion")
        .order_by("-id")[:10]
    )

    # 4) bloqueo por clima de su organización (ya está bien)
    bloqueo = (
        BloqueoClima.objects
        .filter(organization=org, is_blocked=True)
        .order_by("-created_at", "-id")
        .first()
    )

    return render(request, "dashboard/socio.html", {
        "embarcaciones": embarcaciones,
        "ultimos": ultimos,
        "bloqueo": bloqueo,
        "menu_kind": "user",
        "active_nav": "dashboard",
    })


def public_register(request):
    if request.method == 'POST':
        form = PublicRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "¡Gracias! Tu solicitud quedó en revisión. Te avisaremos por correo.")
            return redirect('accounts:login')
    else:
        form = PublicRegisterForm()
    return render(request, 'accounts/public_register.html', {'form': form})

@login_required
@permission_required('accounts.view_applicant', raise_exception=True)
def review_list(request):
    postulantes = Applicant.objects.filter(estado='pending').order_by('-created_at')
    return render(request, 'accounts/review_list.html', {'postulantes': postulantes})

@login_required
@permission_required('accounts.change_applicant', raise_exception=True)
def approve_applicant(request, pk):
    if request.method != 'POST':
        return redirect('accounts:review_list')

    a = get_object_or_404(Applicant, pk=pk)

    # Ya procesada
    if a.estado != 'pending':
        messages.info(request, "Esta postulación ya fue procesada.")
        return redirect('accounts:review_list')

    # Validación robusta de licencia
    lic  = (a.numero_licencia or '').strip()
    venc = a.vencimiento_licencia  # DateField o None
    if not lic or venc is None:
        messages.error(request, "Faltan datos de licencia (número o vencimiento).")
        return redirect('accounts:review_list')

    # Crear/actualizar Usuario
    username = (a.email or a.rut).strip().lower()
    temp_password = get_random_string(10)

    User = get_user_model()
    user, _ = User.objects.get_or_create(
        username=username,
        defaults={
            'email': a.email,
            'first_name': a.nombre,
            'last_name': a.apellido,
        }
    )
    # Sincroniza datos básicos y clave temporal
    user.set_password(temp_password)
    user.email = a.email
    user.first_name = a.nombre
    user.last_name  = a.apellido
    user.save()

    # Organización vía UserProfile
    reviewer_prof = UserProfile.objects.select_related('organization').get(user=request.user)
    up, _ = UserProfile.objects.get_or_create(
        user=user,
        defaults={'organization': reviewer_prof.organization}
    )
    if up.organization_id != reviewer_prof.organization_id:
        up.organization = reviewer_prof.organization
        up.save(update_fields=['organization'])

    # Forzar cambio de clave en UserProfile
    up = UserProfile.objects.get(user=user)
    up.must_change_password = True
    up.save(update_fields=["must_change_password"])

    # (Opcional: si quieres mantener Profile por compatibilidad de datos, lo dejas)
    Profile.objects.update_or_create(
        user=user,
        defaults={
            'must_change_password': True,
            'numero_licencia': lic,
            'vencimiento_licencia': venc,
        }
    )

    # Agregar al grupo "Socio" si existe
    try:
        socio = Group.objects.get(name='Socio')
        user.groups.add(socio)
    except Group.DoesNotExist:
        pass

    # Marcar Applicant aprobado
    a.estado = 'approved'
    a.reviewed_by = request.user
    a.reviewed_at = timezone.now()
    a.save(update_fields=['estado', 'reviewed_by', 'reviewed_at'])

    # Envío de correo (en local saldrá por consola si usas EMAIL_BACKEND de consola)
    try:
        login_url = request.build_absolute_uri(reverse('accounts:login'))
        cuerpo = (
            f"Hola {a.nombre},\n\n"
            f"Tu solicitud fue aprobada.\n\n"
            f"Usuario: {user.username}\n"
            f"Contraseña temporal: {temp_password}\n\n"
            f"Por seguridad, cambia la contraseña al iniciar sesión.\n"
            f"Ingresar: {login_url}\n"
        )
        send_mail("Tu acceso al Club Náutico", cuerpo, None, [a.email], fail_silently=False)
    except Exception:
        # No detenemos el flujo si falla el envío
        pass

    messages.success(request, f"{a.nombre} aprobado.")
    return redirect('accounts:review_list')


@login_required
@permission_required('accounts.change_applicant', raise_exception=True)
def reject_applicant(request, pk):
    if request.method != 'POST':
        return redirect('accounts:review_list')

    Applicant = apps.get_model('accounts', 'Applicant')
    a = get_object_or_404(Applicant, pk=pk)

    if a.estado != 'pending':
        messages.info(request, "Esta postulación ya fue procesada.")
        return redirect('accounts:review_list')

    motivo = (request.POST.get('motivo') or '').strip()

    a.estado = 'rejected'
    a.motivo_revision = motivo
    a.reviewed_by = request.user
    a.reviewed_at = timezone.now()
    a.save(update_fields=['estado', 'motivo_revision', 'reviewed_by', 'reviewed_at'])

    # (Opcional) Notificar por correo al solicitante que fue rechazado
    try:
        cuerpo = (
            f"Hola {a.nombre},\n\n"
            f"Tu solicitud fue rechazada.\n"
            f"{'Motivo: ' + motivo if motivo else ''}\n"
        )
        send_mail("Actualización de solicitud", cuerpo, None, [a.email], fail_silently=True)
    except Exception:
        pass

    messages.warning(request, f"{a.nombre} rechazado.")
    return redirect('accounts:review_list')

class CambiarPasswordView(PasswordChangeView):
    template_name = 'registration/password_change_form.html'
    success_url = reverse_lazy('accounts:password_change_done')

    def form_valid(self, form):
        response = super().form_valid(form)
        # Aquí marcamos que ya no es obligatorio cambiarla
        if hasattr(self.request.user, 'profile') and self.request.user.profile.must_change_password:
            self.request.user.profile.must_change_password = False
            self.request.user.profile.save(update_fields=['must_change_password'])
        return response

class CambioClaveInicialView(PasswordChangeView):
    template_name = 'registration/password_change_form.html'   # <<< nuestro HTML simple
    success_url = reverse_lazy('accounts:dashboard')            # tras guardar, va al dashboard

    def form_valid(self, form):
        resp = super().form_valid(form)
        # bajar el flag para que el middleware no vuelva a forzar cambio
        if hasattr(self.request.user, 'profile'):
            self.request.user.profile.must_change_password = False
            self.request.user.profile.save(update_fields=['must_change_password'])
        return resp

User = get_user_model()

@login_required
def perfil(request):
    profile = UserProfile.objects.select_related("organization").get(user=request.user)
    return render(request, "accounts/perfil.html", {
        "profile": profile,
        "menu_kind": menu_kind_for(request.user),
        "active_nav": "perfil",
    })


class PerfilEditForm(forms.Form):
    email = forms.EmailField(label="Email")
    telefono = forms.CharField(label="Teléfono", max_length=30, required=False)
    foto = forms.ImageField(label="Foto", required=False)

    def __init__(self, *args, **kwargs):
        user = kwargs.pop("user")
        profile = kwargs.pop("profile")
        super().__init__(*args, **kwargs)
        self.user = user
        self.profile = profile
        self.fields["email"].initial = user.email
        self.fields["telefono"].initial = profile.telefono

    def save(self):
        self.user.email = self.cleaned_data["email"]
        self.user.save(update_fields=["email"])
        tel = self.cleaned_data.get("telefono", "")
        self.profile.telefono = tel
        foto = self.cleaned_data.get("foto")
        if foto:
            self.profile.foto = foto
        self.profile.save(update_fields=["telefono", "foto"])


@login_required
def perfil_editar(request):
    profile = UserProfile.objects.get(user=request.user)
    if request.method == "POST":
        form = PerfilEditForm(request.POST, request.FILES, user=request.user, profile=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Datos actualizados.")
            return redirect("accounts:perfil")
    else:
        form = PerfilEditForm(user=request.user, profile=profile)

    return render(request, "accounts/perfil_editar.html", {
        "form": form,
        "profile": profile,
        "menu_kind": menu_kind_for(request.user),
        "active_nav": "perfil",
    })


class UsernameChangeForm(forms.Form):
    username = forms.CharField(label="Nuevo nombre de usuario", max_length=150)

    def __init__(self, *args, **kwargs):
        self.user = kwargs.pop("user")
        self.profile = kwargs.pop("profile")
        super().__init__(*args, **kwargs)

    def clean_username(self):
        new = self.cleaned_data["username"].strip()
        if new == self.user.username:
            raise forms.ValidationError("El usuario nuevo es igual al actual.")
        # único
        if User.objects.filter(username__iexact=new).exists():
            raise forms.ValidationError("Ese nombre de usuario ya está en uso.")
        return new

    def clean(self):
        cleaned = super().clean()
        # regla de 20 días
        if not self.profile.can_change_username():
            raise forms.ValidationError(
                "Aún no puedes cambiar tu usuario. Inténtalo más adelante."
            )
        return cleaned

    def save(self):
        self.user.username = self.cleaned_data["username"]
        self.user.save(update_fields=["username"])
        self.profile.last_username_change = timezone.now()
        self.profile.save(update_fields=["last_username_change"])


@login_required
def perfil_cambiar_usuario(request):
    profile = UserProfile.objects.get(user=request.user)
    if request.method == "POST":
        form = UsernameChangeForm(request.POST, user=request.user, profile=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Nombre de usuario actualizado.")
            return redirect("accounts:perfil")
    else:
        form = UsernameChangeForm(user=request.user, profile=profile)

    # fecha/habilitación opcional si la manejas en el modelo
    can_change = getattr(profile, "can_change_username", lambda: True)()
    last_change = getattr(profile, "last_username_change", None)

    return render(request, "accounts/perfil_cambiar_usuario.html", {
        "form": form,
        "profile": profile,
        "can_change": can_change,
        "last_change": last_change,
        "menu_kind": menu_kind_for(request.user),
        "active_nav": "perfil",
    })

@login_required
def mis_embarcaciones(request):
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    org = prof.organization

    embarcaciones = (
        Embarcacion.objects
        .filter(organization=org, propietario=prof)   # <<< org + dueño
        .select_related("amarra", "organization", "propietario")
    )

    return render(request, "accounts/mis_embarcaciones.html", {
        "embarcaciones": embarcaciones,
        "menu_kind": "user",
        "active_nav": "mis_embarcaciones",
    })

@login_required
@user_passes_test(is_supervisor)
def super_socios(request):
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    org = prof.organization

    socios = (
        UserProfile.objects
        .filter(organization=org)
        .select_related("user", "organization")
        .order_by("user__last_name", "user__first_name")
    )

    return render(request, "accounts/socios_list.html", {
        "socios": socios,
        "menu_kind": menu_kind_for(request.user),
        "active_nav": "socios",
    })

# --- LISTA DE SOCIOS (con filtro simple q=) ---
@login_required
@user_passes_test(is_supervisor)
def socios_list(request):
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    org = prof.organization

    q = request.GET.get("q", "").strip()
    socios = (
        UserProfile.objects
        .select_related("user", "organization")
        .filter(organization=org)
        .order_by("user__last_name", "user__first_name")
    )
    if q:
        socios = socios.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)  |
            Q(user__email__icontains=q)      |
            Q(rut__icontains=q)              |
            Q(telefono__icontains=q)
        )

    return render(request, "accounts/socios_list.html", {
        "socios": socios,
        "q": q,
        # para menú:
        "menu_kind": menu_kind_for(request.user),
        "active_nav": "socios",
    })

@login_required
@user_passes_test(is_supervisor)
def socios_export_xlsx(request):
    # --- mismo filtro simple que usas en socios_list ---
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    org = prof.organization

    q = (request.GET.get("q") or "").strip()
    socios = (
        UserProfile.objects
        .select_related("user", "organization")
        .filter(organization=org)
        .order_by("user__last_name", "user__first_name")
    )
    if q:
        socios = socios.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)  |
            Q(user__email__icontains=q)      |  # << antes: email__icontains
            Q(rut__icontains=q)              |
            Q(telefono__icontains=q)
        )

    # --- armamos el XLSX con formato ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Socios"

    headers = ["Apellido", "Nombre", "Email", "RUT", "Teléfono", "N° Licencia", "Organización"]
    col_widths = [18, 18, 34, 16, 16, 16, 26]

    # Estilos
    head_font  = Font(bold=True, color="FFFFFF")
    head_fill  = PatternFill("solid", fgColor="4F81BD")  # azul
    thin       = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left  = Alignment(horizontal="left", vertical="center", wrap_text=False)
    center= Alignment(horizontal="center", vertical="center")

    # Encabezados
    for c, (text, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=c, value=text)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = cell_border
        ws.column_dimensions[chr(64+c)].width = width

    # Filas
    zebra = PatternFill("solid", fgColor="F7F7F7")
    rownum = 1
    for s in socios:
        rownum += 1
        fila = [
            s.user.last_name or "",
            s.user.first_name or "",
            s.user.email or "",                 # << antes: s.email
            s.rut or "—",
            s.telefono or "—",
            getattr(s, "licencia_numero", None) or "—",
            s.organization.name if s.organization_id else "",
        ]
        for c, val in enumerate(fila, start=1):
            cell = ws.cell(row=rownum, column=c, value=val)
            cell.alignment = left
            cell.border = cell_border
            # rayado
            if rownum % 2 == 0:
                cell.fill = zebra

    # Tabla con estilo (auto-filtro incluido)
    if rownum >= 1:
        tbl = Table(displayName="TblSocios", ref=f"A1:{chr(64+len(headers))}{rownum}")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)

    # Congelar encabezado
    ws.freeze_panes = "A2"

    # Respuesta HTTP
    fecha = timezone.localdate().strftime("%Y%m%d")
    filename = f"Socios_{fecha}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


# --- LISTA DE EMBARCACIONES (con filtros por propietario y tipo) ---
@login_required
@user_passes_test(is_supervisor)
def embarcaciones_list(request):
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    org = prof.organization

    tipo = request.GET.get("tipo", "").strip()      # 'YATE'/'LANCHA'/'MOTO_AGUA'/''...
    owner = request.GET.get("owner", "").strip()    # id de UserProfile (propietario)
    q = request.GET.get("q", "").strip()

    qs = (
        Embarcacion.objects
        .select_related("propietario__user", "amarra", "organization")
        .filter(organization=org)
        .order_by("matricula")
    )

    if tipo:
        qs = qs.filter(tipo=tipo)

    if owner.isdigit():
        qs = qs.filter(propietario_id=int(owner))

    if q:
        qs = qs.filter(
            Q(matricula__icontains=q) |
            Q(propietario__user__first_name__icontains=q) |
            Q(propietario__user__last_name__icontains=q)
        )

    propietarios = UserProfile.objects.filter(organization=org).select_related("user").order_by(
        "user__last_name","user__first_name"
    )

    TIPOS = getattr(Embarcacion, "TIPOS", (
        ("YATE","Yate"), ("LANCHA","Lancha"), ("MOTO_AGUA","Moto de agua"), ("OTRO","Otro")
    ))

    return render(request, "accounts/embarcaciones_list.html", {
        "embarcaciones": qs,
        "propietarios": propietarios,
        "TIPOS": TIPOS,
        "f_tipo": tipo,
        "f_owner": owner,
        "q": q,
        # para menú:
        "menu_kind": menu_kind_for(request.user),
        "active_nav": "embarcaciones",
    })

@login_required
@user_passes_test(is_supervisor)
def contratos_list(request):
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    org = prof.organization

    qs = (ContratoExterno.objects
            .filter(organization=org)
            .select_related("supervisor")
            .order_by("-creado_en"))

    q = request.GET.get("q", "").strip()
    vigencia = request.GET.get("vigencia", "activos")  # activos|inactivos|todos
    estado = request.GET.get("estado", "").strip()     # PENDIENTE|APROBADO|RECHAZADO
    tipo = request.GET.get("tipo", "").strip()         # BOTE|LANCHA|MOTO_AGUA|OTRO

    # --- filtro texto libre
    if q:
        qs = qs.filter(
            Q(nombre__icontains=q) |
            Q(apellido__icontains=q) |
            Q(rut__icontains=q) |
            Q(email__icontains=q) |
            Q(telefono__icontains=q) |
            Q(licencia_numero__icontains=q) |
            Q(detalle_artefacto__icontains=q)
        )

    # --- filtro vigencia por fechas
    hoy = timezone.localdate()
    if vigencia == "activos":
        qs = qs.filter(Q(fecha_fin__isnull=True) | Q(fecha_fin__gte=hoy))
    elif vigencia == "inactivos":
        qs = qs.filter(fecha_fin__lt=hoy)
    # "todos" no filtra nada

    # --- filtro estado (si es válido)
    estados_validos = dict(getattr(ContratoExterno, "ESTADOS", [])).keys()
    if estado and estado in estados_validos:
        qs = qs.filter(estado=estado)

    # --- filtro tipo artefacto (si es válido)
    tipos_validos = dict(getattr(ContratoExterno, "TIPO_ARTEFACTO", [])).keys()
    if tipo and tipo in tipos_validos:
        qs = qs.filter(tipo_artefacto=tipo)

    # --- paginación (25 por página)
    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # preserva filtros para los links de paginación/export
    params = request.GET.copy()
    params.pop("page", None)
    base_qs = params.urlencode()
    page_qs = f"?{base_qs}&" if base_qs else "?"

    return render(request, "accounts/contratos_list.html", {
        "contratos": page_obj.object_list,         # itera con 'contratos' como ya tienes
        "page_obj": page_obj,
        "paginator": paginator,
        "page_qs": page_qs,                        # prefijo para los href de paginación
        "q": q,
        "vigencia": vigencia,
        "estado": estado,
        "tipo": tipo,
        "menu_kind": menu_kind_for(request.user),
        "active_nav": "contratos",
        "ESTADOS": getattr(ContratoExterno, "ESTADOS", []),
        "TIPOS": getattr(ContratoExterno, "TIPO_ARTEFACTO", []),
    })


# EXPORTAR CONTRATOS A EXCEL (respeta los mismos filtros de la lista)
@login_required
@user_passes_test(is_supervisor)
def contratos_export_xlsx(request):
    prof = UserProfile.objects.select_related("organization").get(user=request.user)
    org = prof.organization

    qs = (ContratoExterno.objects
            .filter(organization=org)
            .select_related("supervisor")
            .order_by("-creado_en"))

    q = request.GET.get("q", "").strip()
    vigencia = request.GET.get("vigencia", "activos")
    estado = request.GET.get("estado", "").strip()
    tipo = request.GET.get("tipo", "").strip()

    hoy = timezone.localdate()

    if q:
        qs = qs.filter(
            Q(nombre__icontains=q) |
            Q(apellido__icontains=q) |
            Q(rut__icontains=q) |
            Q(email__icontains=q) |
            Q(telefono__icontains=q) |
            Q(licencia_numero__icontains=q) |
            Q(detalle_artefacto__icontains=q)
        )

    if vigencia == "activos":
        qs = qs.filter(Q(fecha_fin__isnull=True) | Q(fecha_fin__gte=hoy))
    elif vigencia == "inactivos":
        qs = qs.filter(fecha_fin__lt=hoy)

    estados_validos = dict(getattr(ContratoExterno, "ESTADOS", [])).keys()
    if estado and estado in estados_validos:
        qs = qs.filter(estado=estado)

    tipos_validos = dict(getattr(ContratoExterno, "TIPO_ARTEFACTO", [])).keys()
    if tipo and tipo in tipos_validos:
        qs = qs.filter(tipo_artefacto=tipo)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contratos"

    headers = [
        "ID", "Apellido", "Nombre", "RUT", "Email", "Teléfono",
        "Tipo", "Detalle", "Inicio", "Fin", "Vigente",
        "Estado", "Supervisor", "N° Licencia", "Venc. Licencia", "Licencia ok",
        "Observaciones", "Creado", "Actualizado"
    ]

    # Estilos encabezado
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = head_font
        c.fill = head_fill
        c.alignment = center
        c.border = border

    def _vigente(c):
        if not c.fecha_fin:
            return "Sí"
        return "Sí" if c.fecha_fin >= hoy else "No"

    for c in qs:
        supervisor = c.supervisor.get_full_name() if c.supervisor else ""
        row = [
            c.id,
            (c.apellido or ""),
            (c.nombre or ""),
            (c.rut or ""),
            (c.email or ""),
            (c.telefono or ""),
            getattr(c, "get_tipo_artefacto_display", lambda: "")(),
            (c.detalle_artefacto or ""),
            c.fecha_inicio.isoformat() if c.fecha_inicio else "",
            c.fecha_fin.isoformat() if c.fecha_fin else "",
            _vigente(c),
            c.estado if hasattr(c, "estado") else "",
            supervisor,
            (getattr(c, "licencia_numero", "") or ""),
            (c.licencia_vencimiento.isoformat() if getattr(c, "licencia_vencimiento", None) else ""),
            ("Sí" if getattr(c, "licencia_validada", False) else "No"),
            (c.observaciones or "") if hasattr(c, "observaciones") else "",
            c.creado_en.strftime("%Y-%m-%d %H:%M") if hasattr(c, "creado_en") and c.creado_en else "",
            c.actualizado_en.strftime("%Y-%m-%d %H:%M") if hasattr(c, "actualizado_en") and c.actualizado_en else "",
        ]
        ws.append(row)

    # ancho de columnas
    widths = [6, 16, 16, 14, 28, 14, 14, 22, 12, 12, 10, 12, 22, 14, 14, 12, 36, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # bordes suaves al cuerpo
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for c in r:
            c.border = border
            if c.column in (9, 10, 18, 19):  # fechas
                c.alignment = Alignment(horizontal="center")

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="contratos.xlsx"'
    wb.save(resp)
    return resp








